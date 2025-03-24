import os
from openpyxl import Workbook, load_workbook
import processing
excel_file = 'output.xlsx'
headers = ['SR NO', 'DATE', 'VEHICLE NO', 'CONTAINER NO', 'SIZE', 'STATUS', 'FROM', 'TO', 'AMOUNT']

data = processing.invoice_processor.process_invoice()
print(data)


def amount_cal(ContainerNos:list, SwalNo:int) -> tuple['float', 'int']:
    if len(ContainerNos) > 1 and SwalNo == 20:
        return 2000, 2
    else: 
        return 2200, 1

amount, num_rows = amount_cal(data['CONTAINER NO'].split(','), int(data['SIZE']))

if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)

sr_no = len(ws['A'])
container_nos = data['CONTAINER NO'].split(',')
for i in range(num_rows):
    current_container = container_nos[i].strip() if i < len(container_nos) else container_nos[0].strip()
    row = [
        sr_no + i + 1,
        data['DATE'],
        data['VEHICLE NO'],
        current_container,
        data['SIZE'],
        'LOAD' if current_container else 'EMPTY',
        data['FROM'],
        data['TO'],
        amount
    ]

    ws.append(row)

wb.save(excel_file)